"""
Generate HTML Before/After Report for processed product images.
Shows product name, original image, and processed image side by side.
"""

import os
import io
import base64
import requests
import openpyxl
from PIL import Image

# ============================================================
# CONFIG
# ============================================================
FEED_URL = "https://www.engromag.ro/feed/products/ec219e068cd4552bf8759292992425a6"
OUTPUT_DIR = "output"
MAX_PRODUCTS = 200
HTML_FILE = "output/raport_before_after.html"


def load_feed():
    """Download and parse the XLSX feed."""
    print("📥 Descarc feed-ul...")
    resp = requests.get(FEED_URL, timeout=30)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    headers = {}
    for col in range(1, ws.max_column + 1):
        h = str(ws.cell(row=1, column=col).value or "").strip()
        headers[h] = col

    products = []
    for row_idx in range(2, min(2 + MAX_PRODUCTS, ws.max_row + 1)):
        pid = ws.cell(row=row_idx, column=headers.get("ID produs", 1)).value
        sku = ws.cell(row=row_idx, column=headers.get("Cod produs - SKU", 2)).value
        title = str(ws.cell(row=row_idx, column=headers.get("Nume produs", 3)).value or "")
        img_url = str(ws.cell(row=row_idx, column=headers.get("Imagini", 5)).value or "")

        products.append({
            "nr": row_idx - 1,
            "id": pid,
            "sku": sku,
            "title": title,
            "img_url": img_url,
        })
    return products


def find_output_image(nr, sku, pid):
    filename = f"product_{nr:03d}_{sku or pid}.png"
    safe = "".join(c if c.isalnum() or c in ".-_" else "_" for c in filename)
    path = os.path.join(OUTPUT_DIR, safe)
    if os.path.exists(path):
        return path
    return None


def img_to_base64(path_or_url, is_url=False, max_size=300):
    """Convert image to base64 data URI for embedding in HTML."""
    try:
        if is_url:
            resp = requests.get(path_or_url, timeout=10)
            resp.raise_for_status()
            img = Image.open(io.BytesIO(resp.content)).convert("RGB")
        else:
            img = Image.open(path_or_url).convert("RGB")

        img.thumbnail((max_size, max_size), Image.LANCZOS)
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=80)
        b64 = base64.b64encode(buf.getvalue()).decode()
        return f"data:image/jpeg;base64,{b64}", img.size
    except Exception as e:
        return None, (0, 0)


def find_brosa_products(products):
    """Find brosa products that have AI-generated images."""
    import glob
    brosa_files = glob.glob(os.path.join(OUTPUT_DIR, "brosa_[0-9]*.png"))
    brosa_skus = {}
    for f in brosa_files:
        basename = os.path.basename(f)
        # brosa_9991943.png -> sku = 9991943
        if '_original' not in basename:
            sku = basename.replace('brosa_', '').replace('.png', '')
            orig = os.path.join(OUTPUT_DIR, f"brosa_{sku}_original.png")
            brosa_skus[sku] = {'ai': f, 'original': orig if os.path.exists(orig) else None}
    
    result = []
    for p in products:
        sku = str(p['sku'])
        if sku in brosa_skus:
            result.append({**p, 'brosa_ai': brosa_skus[sku]['ai'], 'brosa_orig': brosa_skus[sku]['original']})
            del brosa_skus[sku]
    
    # Also add brosa SKUs not in the first MAX_PRODUCTS
    for sku, paths in brosa_skus.items():
        result.append({
            'nr': 0, 'id': '', 'sku': sku,
            'title': f'Brosa SKU {sku}', 'img_url': '',
            'brosa_ai': paths['ai'], 'brosa_orig': paths['original']
        })
    
    return result


def generate_html(products):
    print(f"\n🌐 Generez HTML cu {len(products)} produse...")

    rows_html = []
    
    # ==========================================
    # BROSE SECTION - PRIMELE SUS
    # ==========================================
    brosa_products = find_brosa_products(products)
    if brosa_products:
        rows_html.append('<div class="section-header">🏵️ BROȘE IMPORTANTE - Procesate cu AI</div>')
        for bp in brosa_products:
            sku = bp['sku']
            title = bp.get('title', f'Brosa {sku}')
            
            # Original from local file
            if bp.get('brosa_orig') and os.path.exists(bp['brosa_orig']):
                orig_b64, orig_size = img_to_base64(bp['brosa_orig'])
            elif bp.get('img_url'):
                orig_b64, orig_size = img_to_base64(bp['img_url'], is_url=True)
            else:
                orig_b64, orig_size = None, (0, 0)
            
            # AI result
            proc_b64, proc_size = img_to_base64(bp['brosa_ai'])
            
            orig_img_html = f'<img src="{orig_b64}" alt="Original">' if orig_b64 else '<div class="no-img">Lipsă</div>'
            proc_img_html = f'<img src="{proc_b64}" alt="AI">' if proc_b64 else '<div class="no-img">Lipsă</div>'
            
            rows_html.append(f"""
            <div class="product-card brosa-card">
                <div class="product-header">
                    <span class="product-nr">🏵️</span>
                    <span class="product-sku">SKU: {sku}</span>
                    <span class="status-ai">🤖 AI Outpaint</span>
                </div>
                <div class="product-title">{title[:80]}{'...' if len(str(title)) > 80 else ''}</div>
                <div class="images-row">
                    <div class="image-box">
                        <div class="image-label">ÎNAINTE</div>
                        {orig_img_html}
                        <div class="image-dims">{orig_size[0]}x{orig_size[1]}</div>
                    </div>
                    <div class="arrow">→</div>
                    <div class="image-box">
                        <div class="image-label">DUPĂ (AI)</div>
                        {proc_img_html}
                        <div class="image-dims">1024x1024</div>
                    </div>
                </div>
            </div>
            """)
            print(f"   [BROSA] {sku} - {title[:40]} - OK")
        
        rows_html.append('<div class="section-header" style="margin-top:30px">📦 Restul produselor</div>')
    
    # ==========================================
    # REGULAR PRODUCTS
    # ==========================================
    for i, prod in enumerate(products):
        nr = prod["nr"]
        sku = prod["sku"]
        title = prod["title"]
        img_url = prod["img_url"]
        pid = prod["id"]

        output_path = find_output_image(nr, sku, pid)

        # Original image
        orig_b64, orig_size = img_to_base64(img_url, is_url=True) if img_url else (None, (0, 0))
        # Processed image
        proc_b64, proc_size = img_to_base64(output_path) if output_path else (None, (0, 0))

        # Determine status
        is_square = orig_size[0] > 0 and abs(orig_size[0] - orig_size[1]) < 10
        if not output_path:
            status_class = "status-error"
            status_text = "❌ Eroare"
        elif is_square:
            status_class = "status-ok"
            status_text = "✅ 1:1 OK"
        else:
            status_class = "status-processed"
            status_text = "🔄 Procesat"

        orig_img_html = f'<img src="{orig_b64}" alt="Original">' if orig_b64 else '<div class="no-img">Fără imagine</div>'
        proc_img_html = f'<img src="{proc_b64}" alt="Procesat">' if proc_b64 else '<div class="no-img">Lipsă</div>'

        rows_html.append(f"""
        <div class="product-card">
            <div class="product-header">
                <span class="product-nr">#{nr}</span>
                <span class="product-sku">SKU: {sku}</span>
                <span class="{status_class}">{status_text}</span>
            </div>
            <div class="product-title">{title[:80]}{'...' if len(title) > 80 else ''}</div>
            <div class="images-row">
                <div class="image-box">
                    <div class="image-label">ÎNAINTE</div>
                    {orig_img_html}
                    <div class="image-dims">{orig_size[0]}x{orig_size[1]}</div>
                </div>
                <div class="arrow">→</div>
                <div class="image-box">
                    <div class="image-label">DUPĂ</div>
                    {proc_img_html}
                    <div class="image-dims">1024x1024</div>
                </div>
            </div>
        </div>
        """)

        print(f"   [{nr}/{MAX_PRODUCTS}] {sku} - OK")

    html = f"""<!DOCTYPE html>
<html lang="ro">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Raport Before/After - Produse Procesate</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif;
            background: #f0f2f5;
            color: #333;
            padding: 20px;
        }}
        .header {{
            text-align: center;
            padding: 30px 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border-radius: 16px;
            margin-bottom: 30px;
            box-shadow: 0 4px 15px rgba(102,126,234,0.4);
        }}
        .header h1 {{ font-size: 28px; margin-bottom: 8px; }}
        .header p {{ opacity: 0.9; font-size: 16px; }}

        .stats {{
            display: flex;
            gap: 15px;
            justify-content: center;
            margin-bottom: 30px;
            flex-wrap: wrap;
        }}
        .stat-box {{
            background: white;
            padding: 15px 25px;
            border-radius: 12px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.08);
            text-align: center;
            min-width: 140px;
        }}
        .stat-box .number {{ font-size: 28px; font-weight: 700; }}
        .stat-box .label {{ font-size: 13px; color: #666; margin-top: 4px; }}
        .stat-box.total .number {{ color: #667eea; }}
        .stat-box.ok .number {{ color: #27ae60; }}
        .stat-box.processed .number {{ color: #3498db; }}
        .stat-box.error .number {{ color: #e74c3c; }}

        .section-header {{
            font-size: 22px;
            font-weight: 700;
            color: #333;
            padding: 15px 0;
            border-bottom: 3px solid #667eea;
            margin-bottom: 20px;
        }}
        .brosa-card {{
            border: 3px solid #e91e63 !important;
            background: linear-gradient(135deg, #fff5f7, #ffffff) !important;
        }}
        .status-ai {{
            background: linear-gradient(135deg, #e91e63, #9c27b0);
            color: white;
            padding: 4px 12px;
            border-radius: 20px;
            font-size: 12px;
            font-weight: 600;
        }}

        .filter-bar {{
            display: flex;
            gap: 10px;
            justify-content: center;
            margin-bottom: 25px;
            flex-wrap: wrap;
        }}
        .filter-btn {{
            padding: 8px 20px;
            border: 2px solid #ddd;
            border-radius: 25px;
            background: white;
            cursor: pointer;
            font-size: 14px;
            font-weight: 500;
            transition: all 0.2s;
        }}
        .filter-btn:hover {{ border-color: #667eea; color: #667eea; }}
        .filter-btn.active {{ background: #667eea; color: white; border-color: #667eea; }}

        .products-grid {{
            display: grid;
            grid-template-columns: repeat(auto-fill, minmax(500px, 1fr));
            gap: 20px;
            max-width: 1400px;
            margin: 0 auto;
        }}

        .product-card {{
            background: white;
            border-radius: 12px;
            padding: 18px;
            box-shadow: 0 2px 8px rgba(0,0,0,0.06);
            transition: transform 0.2s, box-shadow 0.2s;
        }}
        .product-card:hover {{
            transform: translateY(-2px);
            box-shadow: 0 4px 16px rgba(0,0,0,0.12);
        }}

        .product-header {{
            display: flex;
            align-items: center;
            gap: 10px;
            margin-bottom: 8px;
        }}
        .product-nr {{
            font-weight: 700;
            color: #667eea;
            font-size: 14px;
        }}
        .product-sku {{
            font-family: monospace;
            background: #f0f2f5;
            padding: 2px 8px;
            border-radius: 4px;
            font-size: 13px;
        }}
        .product-title {{
            font-size: 14px;
            color: #555;
            margin-bottom: 12px;
            line-height: 1.4;
        }}

        .status-ok {{
            background: #d4edda;
            color: #155724;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
            margin-left: auto;
        }}
        .status-processed {{
            background: #cce5ff;
            color: #004085;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
            margin-left: auto;
        }}
        .status-error {{
            background: #f8d7da;
            color: #721c24;
            padding: 2px 10px;
            border-radius: 12px;
            font-size: 12px;
            font-weight: 600;
            margin-left: auto;
        }}

        .images-row {{
            display: flex;
            align-items: center;
            gap: 10px;
            justify-content: center;
        }}
        .image-box {{
            text-align: center;
            flex: 1;
        }}
        .image-label {{
            font-size: 11px;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 1px;
            color: #888;
            margin-bottom: 6px;
        }}
        .image-box img {{
            max-width: 200px;
            max-height: 200px;
            border-radius: 8px;
            border: 1px solid #eee;
            object-fit: contain;
            background: #fafafa;
        }}
        .image-dims {{
            font-size: 11px;
            color: #aaa;
            margin-top: 4px;
            font-family: monospace;
        }}
        .arrow {{
            font-size: 28px;
            color: #667eea;
            font-weight: 700;
            padding: 0 5px;
        }}
        .no-img {{
            width: 200px;
            height: 200px;
            display: flex;
            align-items: center;
            justify-content: center;
            background: #f5f5f5;
            border-radius: 8px;
            color: #aaa;
            font-size: 14px;
            margin: 0 auto;
        }}

        .search-box {{
            display: block;
            width: 100%;
            max-width: 400px;
            margin: 0 auto 25px;
            padding: 12px 20px;
            border: 2px solid #ddd;
            border-radius: 25px;
            font-size: 15px;
            outline: none;
            transition: border-color 0.2s;
        }}
        .search-box:focus {{ border-color: #667eea; }}

        @media (max-width: 600px) {{
            .products-grid {{ grid-template-columns: 1fr; }}
            .images-row {{ flex-direction: column; }}
            .arrow {{ transform: rotate(90deg); }}
        }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📸 Raport Before / After</h1>
        <p>Produse procesate din feed MerchantPro — {len(products)} produse</p>
    </div>

    <div class="stats">
        <div class="stat-box total">
            <div class="number">{len(products)}</div>
            <div class="label">Total Produse</div>
        </div>
        <div class="stat-box ok">
            <div class="number" id="count-ok">-</div>
            <div class="label">Deja 1:1</div>
        </div>
        <div class="stat-box processed">
            <div class="number" id="count-processed">-</div>
            <div class="label">Procesate</div>
        </div>
        <div class="stat-box error">
            <div class="number" id="count-error">-</div>
            <div class="label">Erori</div>
        </div>
    </div>

    <input type="text" class="search-box" placeholder="🔍 Caută după nume sau SKU..." oninput="filterProducts()">

    <div class="filter-bar">
        <button class="filter-btn active" onclick="setFilter('all', this)">Toate</button>
        <button class="filter-btn" onclick="setFilter('status-ok', this)">✅ 1:1 OK</button>
        <button class="filter-btn" onclick="setFilter('status-processed', this)">🔄 Procesate</button>
        <button class="filter-btn" onclick="setFilter('status-error', this)">❌ Erori</button>
    </div>

    <div class="products-grid">
        {''.join(rows_html)}
    </div>

    <script>
        // Count stats
        document.getElementById('count-ok').textContent = document.querySelectorAll('.status-ok').length;
        document.getElementById('count-processed').textContent = document.querySelectorAll('.status-processed').length;
        document.getElementById('count-error').textContent = document.querySelectorAll('.status-error').length;

        let currentFilter = 'all';

        function setFilter(filter, btn) {{
            currentFilter = filter;
            document.querySelectorAll('.filter-btn').forEach(b => b.classList.remove('active'));
            btn.classList.add('active');
            filterProducts();
        }}

        function filterProducts() {{
            const search = document.querySelector('.search-box').value.toLowerCase();
            document.querySelectorAll('.product-card').forEach(card => {{
                const text = card.textContent.toLowerCase();
                const matchSearch = !search || text.includes(search);
                const matchFilter = currentFilter === 'all' || card.querySelector('.' + currentFilter);
                card.style.display = (matchSearch && matchFilter) ? '' : 'none';
            }});
        }}
    </script>
</body>
</html>"""

    with open(HTML_FILE, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"\n✅ HTML salvat: {os.path.abspath(HTML_FILE)}")


def main():
    print("=" * 60)
    print("🌐 Generator Raport Before/After HTML")
    print("=" * 60)

    products = load_feed()
    generate_html(products)

    print(f"\n{'='*60}")
    print(f"✅ GATA! Deschide în browser:")
    print(f"   {os.path.abspath(HTML_FILE)}")
    print(f"{'='*60}")

if __name__ == "__main__":
    main()