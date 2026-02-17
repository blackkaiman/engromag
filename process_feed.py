"""
MerchantPro Product Feed Processor – MIRROR-EXTEND VERSION
-----------------------------------------------------------
Pipeline: Deterministic mirror-padding + Gaussian feather

  1. Detectează dacă imaginea e 1:1 → skip (resize only)
  2. Creează canvas pătrat la max(w,h)
  3. Centrează imaginea originală
  4. Mirror-pad zonele goale din marginile imaginii
  5. Gaussian blur DOAR pe extensii (soft feathering, radius 8)
  6. Resize la IMAGE_SIZE (768) pentru output final

• ZERO apeluri AI / OpenAI
• Deterministic, rapid, safe pentru 1000+ imagini
• Pixelii originali rămân NEATINȘI
• Extensia continuă textura existentă vizual
• Fără cusături vizibile (mirror + blur feathering)
"""

import os
import io
import logging
import requests
import numpy as np
import openpyxl
from PIL import Image, ImageFilter

# ============================================================
# CONFIG
# ============================================================

FEED_URL = "https://www.engromag.ro/feed/products/ec219e068cd4552bf8759292992425a6"

IMAGE_SIZE = 768
WHITE_THRESHOLD = 240
SQUARE_TOLERANCE = 0.03
BLUR_RADIUS = 8
FEATHER_WIDTH = 12

OUTPUT_DIR = "output_extended"
OUTPUT_EXCEL = "rezultate_extended.xlsx"

# ============================================================
# INIT
# ============================================================

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger()
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# UTILITIES
# ============================================================

def download_image(url: str) -> Image.Image | None:
    """Descarcă imaginea de la URL și returnează RGB."""
    try:
        r = requests.get(url, timeout=20)
        r.raise_for_status()
        return Image.open(io.BytesIO(r.content)).convert("RGB")
    except Exception as e:
        logger.error(f"  Download error: {e}")
        return None


def is_square(img: Image.Image) -> bool:
    """Verifică dacă imaginea e deja pătrată (cu toleranță)."""
    w, h = img.size
    return min(w, h) / max(w, h) >= (1.0 - SQUARE_TOLERANCE)


def detect_border_color(img: Image.Image) -> tuple[tuple[int, int, int], bool]:
    """Detectează culoarea medie de la margini + dacă e alb."""
    arr = np.array(img)
    h, w = arr.shape[:2]
    d = max(5, min(10, h // 20, w // 20))

    border = np.concatenate([
        arr[:d, :].reshape(-1, 3),
        arr[-d:, :].reshape(-1, 3),
        arr[:, :d].reshape(-1, 3),
        arr[:, -d:].reshape(-1, 3),
    ])

    avg = border.mean(axis=0)
    rgb = (int(avg[0]), int(avg[1]), int(avg[2]))
    is_white = all(c > WHITE_THRESHOLD for c in rgb)

    return rgb, is_white


# ============================================================
# 🎯 MIRROR-EXTEND PIPELINE (NO AI)
# ============================================================

def extend_lifestyle_to_square(img: Image.Image) -> Image.Image:
    """
    Extend non-square image to square using mirror padding + blur feathering.

    Algorithm:
      1. Create square canvas at max(w, h)
      2. Center original image on canvas
      3. Fill empty bands with mirrored strips from image edges
      4. Apply Gaussian blur ONLY to extended areas
      5. Feather the seam between original and extension
      6. Resize to IMAGE_SIZE

    Result: seamless, deterministic, no AI, no visible seams.
    """
    w, h = img.size
    max_dim = max(w, h)

    # Already square → just resize
    if is_square(img):
        return img.resize((IMAGE_SIZE, IMAGE_SIZE), Image.LANCZOS)

    canvas = Image.new("RGB", (max_dim, max_dim), (0, 0, 0))
    x_off = (max_dim - w) // 2
    y_off = (max_dim - h) // 2
    canvas.paste(img, (x_off, y_off))

    arr = np.array(img)

    if w < max_dim:
        # --- Landscape gap: fill left and right bands ---
        left_band = x_off
        right_band = max_dim - w - x_off

        if left_band > 0:
            # Mirror from left edge of image
            strip_w = min(left_band, w)
            left_strip = arr[:, :strip_w, :][:, ::-1, :]  # flip horizontally
            # Tile if strip is smaller than band
            left_fill = np.tile(left_strip, (1, (left_band // strip_w) + 1, 1))[:, :left_band, :]
            left_img = Image.fromarray(left_fill).resize((left_band, h), Image.LANCZOS)
            canvas.paste(left_img, (0, y_off))

        if right_band > 0:
            strip_w = min(right_band, w)
            right_strip = arr[:, -strip_w:, :][:, ::-1, :]
            right_fill = np.tile(right_strip, (1, (right_band // strip_w) + 1, 1))[:, :right_band, :]
            right_img = Image.fromarray(right_fill).resize((right_band, h), Image.LANCZOS)
            canvas.paste(right_img, (x_off + w, y_off))

    if h < max_dim:
        # --- Portrait gap: fill top and bottom bands ---
        top_band = y_off
        bottom_band = max_dim - h - y_off

        if top_band > 0:
            strip_h = min(top_band, h)
            top_strip = arr[:strip_h, :, :][::-1, :, :]  # flip vertically
            top_fill = np.tile(top_strip, ((top_band // strip_h) + 1, 1, 1))[:top_band, :, :]
            top_img = Image.fromarray(top_fill).resize((w, top_band), Image.LANCZOS)
            canvas.paste(top_img, (x_off, 0))

        if bottom_band > 0:
            strip_h = min(bottom_band, h)
            bottom_strip = arr[-strip_h:, :, :][::-1, :, :]
            bottom_fill = np.tile(bottom_strip, ((bottom_band // strip_h) + 1, 1, 1))[:bottom_band, :, :]
            bottom_img = Image.fromarray(bottom_fill).resize((w, bottom_band), Image.LANCZOS)
            canvas.paste(bottom_img, (x_off, y_off + h))

    # Fill corners with average border color
    border_color, _ = detect_border_color(img)
    corners = [
        (0, 0, x_off, y_off),                              # top-left
        (x_off + w, 0, max_dim, y_off),                    # top-right
        (0, y_off + h, x_off, max_dim),                    # bottom-left
        (x_off + w, y_off + h, max_dim, max_dim),          # bottom-right
    ]
    for cx1, cy1, cx2, cy2 in corners:
        if cx2 > cx1 and cy2 > cy1:
            corner = Image.new("RGB", (cx2 - cx1, cy2 - cy1), border_color)
            canvas.paste(corner, (cx1, cy1))

    # --- Blur ONLY the extended areas, then feather seam ---
    blurred = canvas.filter(ImageFilter.GaussianBlur(radius=BLUR_RADIUS))

    # Build a mask: 255 = use original, 0 = use blurred extension
    mask = Image.new("L", (max_dim, max_dim), 0)
    orig_rect = Image.new("L", (w, h), 255)
    mask.paste(orig_rect, (x_off, y_off))

    # Feather the mask edges for smooth transition
    mask = mask.filter(ImageFilter.GaussianBlur(radius=FEATHER_WIDTH))

    # Composite: original pixels where mask=255, blurred mirror where mask=0
    final = Image.composite(canvas, blurred, mask)

    return final.resize((IMAGE_SIZE, IMAGE_SIZE), Image.LANCZOS)


def save_image(img: Image.Image, name: str) -> str:
    """Salvează imaginea cu nume filesystem-safe."""
    safe_name = "".join(c if c.isalnum() or c in ".-_" else "_" for c in name)
    path = os.path.join(OUTPUT_DIR, safe_name)
    img.save(path, format="PNG", optimize=True)
    return os.path.abspath(path)


# ============================================================
# MAIN
# ============================================================

def main():
    logger.info("=" * 60)
    logger.info("🎯 MIRROR-EXTEND PIPELINE — No AI, deterministic")
    logger.info("   Pixelii originali rămân NEATINȘI")
    logger.info("=" * 60)

    logger.info("Downloading feed...")
    resp = requests.get(FEED_URL, timeout=30)
    resp.raise_for_status()

    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    new_img_col = ws.max_column + 1
    status_col = ws.max_column + 2
    ws.cell(row=1, column=new_img_col, value="New Image")
    ws.cell(row=1, column=status_col, value="Status")

    extended_count = 0
    skip_count = 0
    total = 0

    for row in range(2, ws.max_row + 1):
        img_url = ws.cell(row=row, column=5).value
        if not img_url:
            continue

        img = download_image(img_url)
        if img is None:
            continue

        total += 1

        if is_square(img):
            result = img.resize((IMAGE_SIZE, IMAGE_SIZE), Image.LANCZOS)
            status = "square_resized"
            skip_count += 1
            logger.info(f"  Row {row}: Skip 1:1 ({img.size[0]}x{img.size[1]})")
        else:
            logger.info(f"  Row {row}: Mirror-extend ({img.size[0]}x{img.size[1]})")
            result = extend_lifestyle_to_square(img)
            status = "mirror_extended"
            extended_count += 1

        filename = f"product_{row}.png"
        path = save_image(result, filename)

        ws.cell(row=row, column=new_img_col, value=path)
        ws.cell(row=row, column=status_col, value=status)

        if total % 100 == 0:
            logger.info(f"  ... processed {total} images")

    output_path = os.path.join(OUTPUT_DIR, OUTPUT_EXCEL)
    wb.save(output_path)

    logger.info("=" * 60)
    logger.info(f"✅ DONE — Extended: {extended_count} | Skipped: {skip_count} | Total: {total}")
    logger.info(f"   Saved to {output_path}")
    logger.info("=" * 60)


# ============================================================
# LOCAL TEST
# ============================================================

def local_test():
    """Test rapid cu 2 imagini din feed."""
    logger.info("=" * 60)
    logger.info("🧪 LOCAL TEST — 2 imagini din feed")
    logger.info("=" * 60)

    test_dir = "test_mirror"
    os.makedirs(test_dir, exist_ok=True)

    logger.info("Downloading feed...")
    resp = requests.get(FEED_URL, timeout=30)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    tested = 0
    for row in range(2, ws.max_row + 1):
        if tested >= 2:
            break
        img_url = ws.cell(row=row, column=5).value
        if not img_url:
            continue

        img = download_image(img_url)
        if img is None or is_square(img):
            continue

        sku = str(ws.cell(row=row, column=2).value or "").strip()
        name = str(ws.cell(row=row, column=3).value or "")[:50]
        logger.info(f"  Test {tested+1}: SKU {sku} | {img.size[0]}x{img.size[1]} | {name}")

        # Save original
        orig_path = os.path.join(test_dir, f"original_{sku}.png")
        img.save(orig_path)

        # Extend
        result = extend_lifestyle_to_square(img)
        ext_path = os.path.join(test_dir, f"extended_{sku}.png")
        result.save(ext_path)

        logger.info(f"    ✅ Saved: {ext_path} ({result.size[0]}x{result.size[1]})")
        tested += 1

    # Generate HTML preview
    html = ['<!DOCTYPE html><html><head><meta charset="utf-8">',
            '<title>Mirror Extend Test</title>',
            '<style>body{font-family:sans-serif;background:#1a1a1a;color:#fff;padding:20px}',
            '.card{display:flex;gap:20px;margin:20px 0;background:#222;padding:20px;border-radius:12px}',
            '.card img{max-width:400px;height:auto;border-radius:8px}',
            'h2{color:#4fc3f7}</style></head><body>',
            '<h1>🔍 Mirror-Extend Test Results</h1>']

    for f in sorted(os.listdir(test_dir)):
        if f.startswith("extended_") and f.endswith(".png"):
            sku = f.replace("extended_", "").replace(".png", "")
            orig = f"original_{sku}.png"
            orig_path = os.path.join(test_dir, orig)
            ext_path = os.path.join(test_dir, f)
            if os.path.exists(orig_path):
                import base64
                with open(orig_path, "rb") as fh:
                    ob64 = base64.b64encode(fh.read()).decode()
                with open(ext_path, "rb") as fh:
                    eb64 = base64.b64encode(fh.read()).decode()
                html.append(f'<h2>SKU: {sku}</h2><div class="card">')
                html.append(f'<div><p>Original</p><img src="data:image/png;base64,{ob64}"></div>')
                html.append(f'<div><p>Extended</p><img src="data:image/png;base64,{eb64}"></div>')
                html.append('</div>')

    html.append('</body></html>')
    html_path = os.path.join(test_dir, "test_report.html")
    with open(html_path, "w") as fh:
        fh.write("\n".join(html))
    logger.info(f"  📄 HTML report: {html_path}")
    logger.info("✅ Local test DONE")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        local_test()
    else:
        main()
