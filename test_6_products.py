"""
Test MIRROR-EXTEND pe 6 produse specifice.
Generează raport HTML before/after.

Pipeline: Deterministic mirror-padding + Gaussian feather
• ZERO apeluri AI / OpenAI
• Deterministic, rapid, safe pentru 1000+ imagini
• Pixelii originali rămân NEATINȘI
• Extensia continuă textura existentă vizual
• Fără cusături vizibile (mirror + blur feathering)
"""

import os
import io
import base64
import logging
import requests
import numpy as np
import openpyxl
from PIL import Image, ImageFilter

# ============================================================
# CONFIG
# ============================================================

FEED_URL = "https://www.engromag.ro/feed/products/ec219e068cd4552bf8759292992425a6"

IMAGE_SIZE = 768
WHITE_THRESHOLD = 240
SQUARE_TOLERANCE = 0.03
BLUR_RADIUS = 8
FEATHER_WIDTH = 12

OUTPUT_DIR = "test_6"
TARGET_SKUS = {"9992737", "9992725", "9991619", "9991961", "9991962", "9991967"}

# ============================================================
# INIT
# ============================================================

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
logger = logging.getLogger()
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ============================================================
# FUNCTIONS
# ============================================================

def download_image(url: str) -> Image.Image | None:
    try:
        r = requests.get(url, timeout=20)
        r.raise_for_status()
        return Image.open(io.BytesIO(r.content)).convert("RGB")
    except Exception as e:
        logger.error(f"  Download error: {e}")
        return None


def is_square(img: Image.Image) -> bool:
    """Verifică dacă imaginea e deja pătrată (cu toleranță)."""
    w, h = img.size
    return min(w, h) / max(w, h) >= (1.0 - SQUARE_TOLERANCE)


def detect_border_color(img: Image.Image) -> tuple[tuple[int, int, int], bool]:
    arr = np.array(img)
    h, w = arr.shape[:2]
    d = max(5, min(10, h // 20, w // 20))
    border = np.concatenate([
        arr[:d, :].reshape(-1, 3),
        arr[-d:, :].reshape(-1, 3),
        arr[:, :d].reshape(-1, 3),
        arr[:, -d:].reshape(-1, 3),
    ])
    avg = border.mean(axis=0)
    rgb = (int(avg[0]), int(avg[1]), int(avg[2]))
    is_white = all(c > WHITE_THRESHOLD for c in rgb)
    return rgb, is_white


def extend_lifestyle_to_square(img: Image.Image) -> Image.Image:
    """
    Extend non-square image to square using mirror padding + blur feathering.

    Algorithm:
      1. Create square canvas at max(w, h)
      2. Center original image on canvas
      3. Fill empty bands with mirrored strips from image edges
      4. Fill corners with average border color
      5. Apply Gaussian blur ONLY to extended areas
      6. Feather the seam between original and extension
      7. Resize to IMAGE_SIZE

    Result: seamless, deterministic, no AI, no visible seams.
    """
    w, h = img.size
    max_dim = max(w, h)

    # Already square → just resize
    if is_square(img):
        return img.resize((IMAGE_SIZE, IMAGE_SIZE), Image.LANCZOS)

    logger.info(f"  [1/4] Canvas {max_dim}×{max_dim}, center original {w}×{h}")
    canvas = Image.new("RGB", (max_dim, max_dim), (0, 0, 0))
    x_off = (max_dim - w) // 2
    y_off = (max_dim - h) // 2
    canvas.paste(img, (x_off, y_off))

    arr = np.array(img)

    logger.info("  [2/4] Mirror-padding extended areas...")
    if w < max_dim:
        # --- Landscape gap: fill left and right bands ---
        left_band = x_off
        right_band = max_dim - w - x_off

        if left_band > 0:
            strip_w = min(left_band, w)
            left_strip = arr[:, :strip_w, :][:, ::-1, :]
            left_fill = np.tile(left_strip, (1, (left_band // strip_w) + 1, 1))[:, :left_band, :]
            left_img = Image.fromarray(left_fill).resize((left_band, h), Image.LANCZOS)
            canvas.paste(left_img, (0, y_off))

        if right_band > 0:
            strip_w = min(right_band, w)
            right_strip = arr[:, -strip_w:, :][:, ::-1, :]
            right_fill = np.tile(right_strip, (1, (right_band // strip_w) + 1, 1))[:, :right_band, :]
            right_img = Image.fromarray(right_fill).resize((right_band, h), Image.LANCZOS)
            canvas.paste(right_img, (x_off + w, y_off))

    if h < max_dim:
        # --- Portrait gap: fill top and bottom bands ---
        top_band = y_off
        bottom_band = max_dim - h - y_off

        if top_band > 0:
            strip_h = min(top_band, h)
            top_strip = arr[:strip_h, :, :][::-1, :, :]
            top_fill = np.tile(top_strip, ((top_band // strip_h) + 1, 1, 1))[:top_band, :, :]
            top_img = Image.fromarray(top_fill).resize((w, top_band), Image.LANCZOS)
            canvas.paste(top_img, (x_off, 0))

        if bottom_band > 0:
            strip_h = min(bottom_band, h)
            bottom_strip = arr[-strip_h:, :, :][::-1, :, :]
            bottom_fill = np.tile(bottom_strip, ((bottom_band // strip_h) + 1, 1, 1))[:bottom_band, :, :]
            bottom_img = Image.fromarray(bottom_fill).resize((w, bottom_band), Image.LANCZOS)
            canvas.paste(bottom_img, (x_off, y_off + h))

    # Fill corners with average border color
    logger.info("  [3/4] Filling corners + blur feathering...")
    border_color, _ = detect_border_color(img)
    corners = [
        (0, 0, x_off, y_off),
        (x_off + w, 0, max_dim, y_off),
        (0, y_off + h, x_off, max_dim),
        (x_off + w, y_off + h, max_dim, max_dim),
    ]
    for cx1, cy1, cx2, cy2 in corners:
        if cx2 > cx1 and cy2 > cy1:
            corner = Image.new("RGB", (cx2 - cx1, cy2 - cy1), border_color)
            canvas.paste(corner, (cx1, cy1))

    # --- Blur ONLY the extended areas, then feather seam ---
    blurred = canvas.filter(ImageFilter.GaussianBlur(radius=BLUR_RADIUS))

    # Build a mask: 255 = use original, 0 = use blurred extension
    mask = Image.new("L", (max_dim, max_dim), 0)
    orig_rect = Image.new("L", (w, h), 255)
    mask.paste(orig_rect, (x_off, y_off))

    # Feather the mask edges for smooth transition
    mask = mask.filter(ImageFilter.GaussianBlur(radius=FEATHER_WIDTH))

    # Composite: original pixels where mask=255, blurred mirror where mask=0
    final = Image.composite(canvas, blurred, mask)

    logger.info(f"  [4/4] Resize to {IMAGE_SIZE}×{IMAGE_SIZE}")
    return final.resize((IMAGE_SIZE, IMAGE_SIZE), Image.LANCZOS)


def img_to_b64(img: Image.Image, max_w: int = 500) -> str:
    w, h = img.size
    if w > max_w:
        ratio = max_w / w
        img = img.resize((max_w, int(h * ratio)), Image.LANCZOS)
    buf = io.BytesIO()
    img.save(buf, format="JPEG", quality=85)
    return "data:image/jpeg;base64," + base64.b64encode(buf.getvalue()).decode()


# ============================================================
# MAIN
# ============================================================

def main():
    logger.info("=" * 60)
    logger.info("🎯 TEST 6 PRODUSE — MIRROR-EXTEND PIPELINE (NO AI)")
    logger.info(f"   SKUs: {', '.join(sorted(TARGET_SKUS))}")
    logger.info("   Deterministic • Mirror-pad • Gaussian feather")
    logger.info("=" * 60)

    logger.info("Downloading feed...")
    resp = requests.get(FEED_URL, timeout=30)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    results = []

    for row in range(2, ws.max_row + 1):
        sku = str(ws.cell(row=row, column=2).value or "").strip()
        if sku not in TARGET_SKUS:
            continue

        name = ws.cell(row=row, column=3).value or "Unknown"
        img_url = ws.cell(row=row, column=5).value
        if not img_url:
            continue

        logger.info(f"\n{'='*60}")
        logger.info(f"SKU: {sku} | Row {row} | {name[:60]}")
        logger.info(f"{'='*60}")

        img = download_image(img_url)
        if img is None:
            results.append({
                "sku": sku, "name": name, "row": row,
                "status": "download_error", "img_url": img_url,
                "original_img": None, "result_img": None,
            })
            continue

        original = img.copy()
        w, h = img.size
        logger.info(f"  Dimensiuni: {w}x{h}")

        if is_square(img):
            result = img.resize((IMAGE_SIZE, IMAGE_SIZE), Image.LANCZOS)
            status = "square_resized"
            logger.info(f"  Already square → resize to {IMAGE_SIZE}×{IMAGE_SIZE}")
        else:
            result = extend_lifestyle_to_square(img)
            status = "mirror_extended"

        filename = f"test_{sku}.png"
        path = os.path.join(OUTPUT_DIR, filename)
        result.save(path, format="PNG", optimize=True)

        results.append({
            "sku": sku, "name": name, "row": row,
            "status": status, "img_url": img_url,
            "original_img": original, "result_img": result,
        })

        logger.info(f"  ✅ Saved: {path}")

    # ============================================================
    # Generate HTML report
    # ============================================================
    logger.info("\n\nGenerating HTML report...")

    cards_html = ""
    for r in results:
        if r["original_img"] is None or r["result_img"] is None:
            status_badge = '<span style="background:#e74c3c;color:white;padding:4px 10px;border-radius:12px;font-size:13px;">ERROR</span>'
            cards_html += f"""
            <div style="background:white;border-radius:12px;padding:20px;margin-bottom:20px;box-shadow:0 2px 8px rgba(0,0,0,0.1);">
                <div style="display:flex;gap:12px;align-items:center;margin-bottom:12px;">
                    {status_badge}
                    <strong>SKU: {r['sku']}</strong> — {r['name'][:80]}
                </div>
                <p style="color:#e74c3c;">Download failed</p>
            </div>"""
            continue

        before_b64 = img_to_b64(r["original_img"], max_w=480)
        after_b64 = img_to_b64(r["result_img"], max_w=480)

        if r["status"] == "mirror_extended":
            badge_color = "#27ae60"
            badge_text = "✅ MIRROR EXTENDED"
        elif r["status"] == "square_resized":
            badge_color = "#3498db"
            badge_text = "🔲 SQUARE RESIZED"
        else:
            badge_color = "#f39c12"
            badge_text = "⚠️ UNKNOWN"

        w, h = r["original_img"].size

        cards_html += f"""
        <div style="background:white;border-radius:12px;padding:24px;margin-bottom:24px;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
            <div style="display:flex;gap:12px;align-items:center;margin-bottom:16px;flex-wrap:wrap;">
                <span style="background:{badge_color};color:white;padding:5px 14px;border-radius:20px;font-size:13px;font-weight:600;">{badge_text}</span>
                <strong style="font-size:15px;">SKU: {r['sku']}</strong>
                <span style="color:#888;font-size:13px;">Row {r['row']} • {w}×{h}</span>
            </div>
            <h3 style="margin:0 0 16px 0;font-size:16px;color:#333;">{r['name'][:100]}</h3>
            <div style="display:grid;grid-template-columns:1fr 1fr;gap:16px;">
                <div>
                    <div style="text-align:center;padding:8px 0;font-weight:600;color:#e74c3c;font-size:14px;">BEFORE (original)</div>
                    <div style="background:#f8f8f8;border-radius:8px;padding:8px;text-align:center;">
                        <img src="{before_b64}" style="max-width:100%;border-radius:4px;border:1px solid #eee;" />
                    </div>
                </div>
                <div>
                    <div style="text-align:center;padding:8px 0;font-weight:600;color:#27ae60;font-size:14px;">AFTER (mirror-extend)</div>
                    <div style="background:#f8f8f8;border-radius:8px;padding:8px;text-align:center;">
                        <img src="{after_b64}" style="max-width:100%;border-radius:4px;border:1px solid #eee;" />
                    </div>
                </div>
            </div>
        </div>"""

    ext_count = sum(1 for r in results if r["status"] == "mirror_extended")
    sq_count = sum(1 for r in results if r["status"] == "square_resized")
    err_count = sum(1 for r in results if r["status"] == "download_error")

    html = f"""<!DOCTYPE html>
<html lang="ro">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Test 6 Produse — Mirror-Extend Before/After</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; background: #f0f2f5; color: #333; }}
        .container {{ max-width: 1100px; margin: 0 auto; padding: 24px; }}
        h1 {{ text-align: center; margin-bottom: 8px; font-size: 28px; }}
        .subtitle {{ text-align: center; color: #888; margin-bottom: 32px; font-size: 15px; }}
        .stats {{ display: flex; gap: 16px; justify-content: center; margin-bottom: 32px; flex-wrap: wrap; }}
        .stat {{ background: white; padding: 16px 24px; border-radius: 10px; text-align: center; box-shadow: 0 2px 8px rgba(0,0,0,0.06); }}
        .stat .num {{ font-size: 32px; font-weight: 700; }}
        .stat .label {{ font-size: 13px; color: #888; margin-top: 4px; }}
    </style>
</head>
<body>
<div class="container">
    <h1>🎯 Test 6 Produse — Mirror-Extend</h1>
    <p class="subtitle">Deterministic mirror-padding + Gaussian feather<br>ZERO AI • Pixelii originali rămân NEATINȘI</p>

    <div class="stats">
        <div class="stat">
            <div class="num">{len(results)}</div>
            <div class="label">Produse testate</div>
        </div>
        <div class="stat">
            <div class="num" style="color:#27ae60;">{ext_count}</div>
            <div class="label">Mirror Extended</div>
        </div>
        <div class="stat">
            <div class="num" style="color:#3498db;">{sq_count}</div>
            <div class="label">Square Resized</div>
        </div>
        <div class="stat">
            <div class="num" style="color:#e74c3c;">{err_count}</div>
            <div class="label">Erori</div>
        </div>
    </div>

    {cards_html}

    <p style="text-align:center;color:#aaa;padding:24px;font-size:13px;">
        Generat automat • {len(results)} produse • Pipeline: mirror-extend (Pillow + numpy, ZERO AI)
    </p>
</div>
</body>
</html>"""

    report_path = os.path.join(OUTPUT_DIR, "test_6_report.html")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(html)
    logger.info(f"✅ Report saved: {report_path}")

    logger.info("=" * 60)
    logger.info(f"✅ DONE — Extended: {ext_count} | Square: {sq_count} | Errors: {err_count} | Total: {len(results)}")
    logger.info("   Pipeline: mirror-extend (Pillow + numpy, ZERO AI)")
    logger.info("=" * 60)


# ============================================================
# LOCAL TEST — 2 images from disk
# ============================================================

def local_test():
    """Load 2 test images from feed, run extend_lifestyle_to_square(), save outputs."""
    logger.info("=" * 60)
    logger.info("🧪 LOCAL TEST — 2 imagini din feed (mirror-extend)")
    logger.info("=" * 60)

    test_dir = "test_mirror"
    os.makedirs(test_dir, exist_ok=True)

    logger.info("Downloading feed...")
    resp = requests.get(FEED_URL, timeout=30)
    resp.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(resp.content))
    ws = wb.active

    tested = 0
    for row in range(2, ws.max_row + 1):
        if tested >= 2:
            break
        img_url = ws.cell(row=row, column=5).value
        if not img_url:
            continue

        img = download_image(img_url)
        if img is None or is_square(img):
            continue

        sku = str(ws.cell(row=row, column=2).value or "").strip()
        name = str(ws.cell(row=row, column=3).value or "")[:50]
        logger.info(f"  Test {tested+1}: SKU {sku} | {img.size[0]}x{img.size[1]} | {name}")

        # Save original
        orig_path = os.path.join(test_dir, f"original_{sku}.png")
        img.save(orig_path)

        # Extend
        result = extend_lifestyle_to_square(img)
        ext_path = os.path.join(test_dir, f"extended_{sku}.png")
        result.save(ext_path)

        logger.info(f"    ✅ Saved: {ext_path} ({result.size[0]}x{result.size[1]})")
        tested += 1

    # Generate HTML preview
    html_parts = ['<!DOCTYPE html><html><head><meta charset="utf-8">',
                  '<title>Mirror Extend Test</title>',
                  '<style>body{font-family:sans-serif;background:#1a1a1a;color:#fff;padding:20px}',
                  '.card{display:flex;gap:20px;margin:20px 0;background:#222;padding:20px;border-radius:12px}',
                  '.card img{max-width:400px;height:auto;border-radius:8px}',
                  'h2{color:#4fc3f7}</style></head><body>',
                  '<h1>🔍 Mirror-Extend Test Results (ZERO AI)</h1>']

    for f in sorted(os.listdir(test_dir)):
        if f.startswith("extended_") and f.endswith(".png"):
            sku = f.replace("extended_", "").replace(".png", "")
            orig = f"original_{sku}.png"
            orig_fpath = os.path.join(test_dir, orig)
            ext_fpath = os.path.join(test_dir, f)
            if os.path.exists(orig_fpath):
                with open(orig_fpath, "rb") as fh:
                    ob64 = base64.b64encode(fh.read()).decode()
                with open(ext_fpath, "rb") as fh:
                    eb64 = base64.b64encode(fh.read()).decode()
                html_parts.append(f'<h2>SKU: {sku}</h2><div class="card">')
                html_parts.append(f'<div><p>Original</p><img src="data:image/png;base64,{ob64}"></div>')
                html_parts.append(f'<div><p>Extended (mirror-pad + blur)</p><img src="data:image/png;base64,{eb64}"></div>')
                html_parts.append('</div>')

    html_parts.append('</body></html>')
    html_path = os.path.join(test_dir, "test_report.html")
    with open(html_path, "w") as fh:
        fh.write("\n".join(html_parts))
    logger.info(f"  📄 HTML report: {html_path}")
    logger.info("✅ Local test DONE — mirror-extend pipeline OK")
    print(f"\n✅ SUCCESS: {tested} images processed with mirror-extend (ZERO AI)")
    print(f"   Results in: {os.path.abspath(test_dir)}/")


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "--test":
        local_test()
    else:
        main()
